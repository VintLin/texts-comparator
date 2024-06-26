import os
import json
import openpyxl

from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

class FileUtils:
    
    @staticmethod
    def read_from_excel(filename):
        workbook = openpyxl.load_workbook(filename)
        items = []
        
        worksheet = workbook.active
        
        for row in list(worksheet.iter_rows(values_only=True)):
            text = row[0]
            if text is not None and len(str(text).strip()) > 0:
                items.append(text)

        workbook.close()
    
    @staticmethod
    def read_from_text(filename):
        with open(filename, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            lines = [line.strip() for line in lines]
        return lines
    
    @staticmethod
    def save_to_json(data, file_path):
        current_time = datetime.now()
        
        formatted_timestamp = current_time.strftime("%Y%m%d")
        
        json_path = os.path.join(file_path, f"CompareResult_{formatted_timestamp}.json")
        
        os.makedirs(os.path.dirname(json_path), exist_ok=True)
        
        with open(json_path, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file)
    
    @staticmethod
    def save_to_execel(data, file_path):
        try:
            current_time = datetime.now()
            
            formatted_timestamp = current_time.strftime("%Y%m%d")
            
            excel_path = os.path.join(file_path, f"CompareResult_{formatted_timestamp}.xlsx")
            
            if os.path.exists(excel_path):
                os.remove(excel_path)
                
            os.makedirs(os.path.dirname(excel_path), exist_ok=True)
            
            try:
                workbook = load_workbook(excel_path)
            except FileNotFoundError:
                workbook = Workbook()
                
            # 获取第一个工作表
            sheet = workbook.active
            head_height = 30
            count_width = 10
            
            current_row = 1
            
            FileUtils._set_sheet_value("Left", current_row, 1, sheet, height=head_height, bold=True, align_horizontal="center", background="FABF8F")
            start_cell = get_column_letter(1) + str(current_row)
            end_cell = get_column_letter(2) + str(current_row)
            sheet.merge_cells(f'{start_cell}:{end_cell}')
            
            FileUtils._set_sheet_value("Right", current_row, 3, sheet, height=head_height, bold=True, align_horizontal="center", background="FABF8F")
            start_cell = get_column_letter(3) + str(current_row)
            end_cell = get_column_letter(4) + str(current_row)
            sheet.merge_cells(f'{start_cell}:{end_cell}')
            current_row += 1
                
            for item in data.values():
                left, rights = item["left"], item["right"]
                right_count = len(rights) if (len(rights) != 0) else 1
                
                start_row = current_row
                end_row = current_row + right_count - 1
                
                start_cell = get_column_letter(1) + str(start_row)
                end_cell = get_column_letter(1) + str(end_row)
                sheet.merge_cells(f'{start_cell}:{end_cell}')
                
                FileUtils._set_sheet_rich_value(FileUtils._get_red_font(left['text'], left['red_marks'], left['tag']), start_row, 1, sheet)
                for index, item in enumerate(rights):
                    is_right = len(item['red_marks']) == 0
                    FileUtils._set_sheet_value("", start_row + index, 2, sheet)
                    FileUtils._set_sheet_rich_value(FileUtils._get_red_font(item['text'], item['red_marks'], item['tag']), start_row + index, 3, sheet)
                    FileUtils._set_sheet_value("Right" if is_right else f"Error\n（{len(item['red_marks'])}）", start_row + index, 4, sheet, width=count_width, bold=True, color="00B050" if is_right else "FF0000")
                
                start_cell = get_column_letter(2) + str(start_row)
                end_cell = get_column_letter(2) + str(end_row)
                sheet.merge_cells(f'{start_cell}:{end_cell}')
                
                is_right = len(left['red_marks']) == 0
                FileUtils._set_sheet_value("Right" if is_right else f"Error\n（{len(left['red_marks'])}）", start_row, 2, sheet, width=count_width, bold=True, color="00B050" if is_right else "FF0000")
                
                current_row += right_count
            
            workbook.save(excel_path)
            
            workbook.close()
            return True
        except:        
            return False
    
    @staticmethod
    def _set_sheet_value(value, row_index: int, column_index: int, sheet, bold=False, width=60, height=60, color="000000", background="ffffff", align_horizontal='left'):
        font = Font(name='微软雅黑', size=11, bold=bold, color=color)
        alignment = Alignment(horizontal=align_horizontal, vertical='center', wrap_text=True)
        cell = sheet.cell(row=row_index, column=column_index, value=value)
        fill = PatternFill(fill_type="solid", fgColor=background)
        border = Border(left=Side(border_style="thin"), 
                        right=Side(border_style="thin"), 
                        top=Side(border_style="thin"), 
                        bottom=Side(border_style="thin"))
        cell.font = font
        cell.alignment = alignment
        cell.fill = fill
        cell.border = border
        sheet.row_dimensions[row_index].height = height
        sheet.column_dimensions[sheet.cell(row=row_index, column=column_index).column_letter].width = width

    @staticmethod
    def _set_sheet_rich_value(value: CellRichText, row_index: int, column_index: int, sheet, background="ffffff"):
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell = sheet.cell(row=row_index, column=column_index, value=value)
        fill = PatternFill(fill_type="solid", fgColor=background)
        border = Border(left=Side(border_style="thin"), 
                        right=Side(border_style="thin"), 
                        top=Side(border_style="thin"), 
                        bottom=Side(border_style="thin"))
        cell.alignment = alignment
        cell.fill = fill
        cell.border = border
        sheet.row_dimensions[row_index].height = 100
        sheet.column_dimensions[sheet.cell(row=row_index, column=column_index).column_letter].width = 60
    
    @staticmethod
    def _get_red_font(text, red_marks, tag = ""):
        red_font = InlineFont(color="FF0000", sz=12, b=True)
        black_font = InlineFont(color="000000", sz=12)
        grey_font = InlineFont(color="888888", sz=12, b=True)
        
        sorted_indices = sorted(set(red_marks))
        rich_string = []
        last_pos = 0
        for index in sorted_indices:
            if index >= len(text):
                continue
            if index > last_pos:
                rich_string.append(TextBlock(black_font, text[last_pos:index]))
            rich_string.append(TextBlock(red_font, text[index]))
            last_pos = index + 1

        if last_pos < len(text):
            rich_string.append(TextBlock(black_font, text[last_pos:]))
        if tag:
            rich_string.append(TextBlock(grey_font, f"\n({tag})"))
        return CellRichText(rich_string)